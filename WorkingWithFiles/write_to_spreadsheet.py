import os
import sys
import xlsxwriter
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

class WriteToSpredSheet:
	

	ROOT_PATH_OF_OUTPUT = ""
	OUTPUT_FILE = ""
	OUTPUT_FILE_PATH = ""
	WORK_BOOK = ""
	WORK_SHEET = ""

	def __init__(self, root_path: str, filename_of_output_file: str) -> bool:
		super(WriteToSpredSheet, self).__init__()
		self.ROOT_PATH_OF_OUTPUT = root_path
		self.OUTPUT_FILE = filename_of_output_file
		self.OUTPUT_FILE_PATH = f"{self.ROOT_PATH_OF_OUTPUT}/{self.OUTPUT_FILE}"
		self.WORK_BOOK = self.create_workbook()
		self.WORK_SHEET = self.create_work_sheet()
		self.add_headers_to_spreadsheet()



	def create_workbook(self):
		try:
			work_book = xlsxwriter.Workbook(self.OUTPUT_FILE_PATH)
			return work_book
		except Exception as e:
			print(e)
			return False

	def create_work_sheet(self):
		try:
			work_sheet = self.WORK_BOOK.add_worksheet()
			self.close_workbook()
			return work_sheet
		except Exception as e:
			print(e)
			return False

	def load_spread_sheet(self):
		self.WORK_BOOK = load_workbook(self.OUTPUT_FILE_PATH)
		self.WORK_SHEET = self.WORK_BOOK.worksheets[0]

	def add_headers_to_spreadsheet(self):
		self.load_spread_sheet()
		col = 1
		try:
			data = [
				"Домен сайта", 
				"Всего пространства", 
				"Свободное пространство", 
				"Занятое пространство", 
				"Веб файлы", 
				"Почта", 
				"База данных", 
				"Конфигурационные файлы",
				"Папа FTP"
			]

			font = Font(
				name='Arial',
				size=14,
				bold=True,
				italic=False,
				vertAlign=None,
				underline='none',
				strike=False,
				color='FF000000'
			)

			for item in data:
				self.WORK_SHEET.cell(row=1, column=col, value=item)
				col += 1

			self.WORK_BOOK.save(self.OUTPUT_FILE_PATH)



			for row in self.WORK_SHEET.iter_rows(1,1):
				for column_index, column in enumerate(row,1):
					column.font = font
					self.WORK_SHEET.column_dimensions[get_column_letter(column_index)].width = 40

			self.WORK_BOOK.save(self.OUTPUT_FILE_PATH)
		except Exception as e:
			print(e)
			return False

	def write_to_spredsheet(self, data: list, row: int, col: int) -> bool:
		try:
			self.load_spread_sheet()

			for item in data:
				self.WORK_SHEET.cell(row=row, column=col, value=item)
				col += 1
			self.WORK_BOOK.save(self.OUTPUT_FILE_PATH)
		except Exception as e:
			print(e)
			return False
		

	def close_workbook(self):
		self.WORK_BOOK.close()












